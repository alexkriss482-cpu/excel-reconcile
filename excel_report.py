"""Формирование Excel-отчёта сверки."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import REPORT_RULES
from reconcile import ReconcileResult

TITLE_FILL = PatternFill(fill_type="solid", fgColor="FF1F4E78")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFFFF")
HEADER_FONT = Font(bold=True)
SUBTOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
RULES_HEADER_FONT = Font(bold=True, size=11, color="FFFFFFFF")
RED_BOLD_FONT = Font(color="FFC00000", bold=True)
DEFAULT_FONT = Font(color="FF000000")
AMOUNT_FORMAT = "#,##0.00"
INTEGER_FORMAT = "#,##0"
DATE_FORMAT = "DD.MM.YYYY"

AMOUNT_HEADERS = frozenset(
    {
        "ОФД безнал",
        "Реестр карт",
        "Разница безнал",
        "Разница",
        "ОФД наличные",
        "Реестр кассовых",
        "Разница наличные",
        "ОФД кредит",
    }
)
INTEGER_HEADERS = frozenset({"Чеков приход", "Возвратов"})

DAILY_HEADERS = [
    "Подразделение",
    "Дата",
    "ОФД безнал",
    "Реестр карт",
    "Разница безнал",
    "Статус безнал",
    "ОФД наличные",
    "Реестр кассовых",
    "Разница наличные",
    "Статус наличные",
    "Чеков приход",
    "Возвратов",
    "ОФД кредит",
    "Источник ОФД",
]

TOTALS_HEADERS = [
    "Подразделение",
    "ОФД безнал",
    "Реестр карт",
    "Разница",
    "Статус",
    "ОФД наличные",
    "Реестр кассовых",
    "Разница",
    "Статус",
    "ОФД кредит",
    "Общий статус",
]

DISC_HEADERS = [
    "Подразделение",
    "Дата",
    "ОФД безнал",
    "Реестр карт",
    "Разница безнал",
    "ОФД наличные",
    "Реестр кассовых",
    "Разница наличные",
    "Источник ОФД",
]


def _write_sheet_title(ws, title: str, width: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_header_row(ws, headers: list[str], row: int = 2) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT


def _apply_discrepancy_formatting(
    ws,
    cashless_range: str,
    cashless_diff_col: str,
    cash_range: str,
    cash_diff_col: str,
) -> None:
    """Красный жирный шрифт для блоков с расхождением — как в эталоне."""
    cashless_rule = FormulaRule(
        formula=[f"ABS({cashless_diff_col}3)>=0.01"],
        font=RED_BOLD_FONT,
    )
    cash_rule = FormulaRule(
        formula=[f"ABS({cash_diff_col}3)>=0.01"],
        font=RED_BOLD_FONT,
    )
    ws.conditional_formatting.add(cashless_range, cashless_rule)
    ws.conditional_formatting.add(cash_range, cash_rule)


def _apply_overall_status_formatting(ws, status_range: str) -> None:
    rule = FormulaRule(
        formula=['$K3="Требует проверки"'],
        font=RED_BOLD_FONT,
    )
    ws.conditional_formatting.add(status_range, rule)


def _apply_cell_format(cell, header: str) -> None:
    """Формат даты, сумм с разрядностью и целых чисел."""
    value = cell.value
    if value is None:
        return
    if header == "Дата":
        cell.number_format = DATE_FORMAT
    elif header in AMOUNT_HEADERS and isinstance(value, (int, float)):
        cell.number_format = AMOUNT_FORMAT
    elif header in INTEGER_HEADERS and isinstance(value, (int, float)):
        cell.number_format = INTEGER_FORMAT


def _autosize(ws, max_col: int, max_row: int) -> None:
    for col in range(1, max_col + 1):
        values = []
        for row in range(1, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                values.append(str(value))
        width = min(max((len(value) for value in values), default=10) + 2, 45)
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_daily_sheet(ws, daily: pd.DataFrame) -> None:
    _write_sheet_title(
        ws,
        "Сверка ОФД с реестрами — обновлено дополнительным отчетом",
        len(DAILY_HEADERS),
    )
    _write_header_row(ws, DAILY_HEADERS)

    for row_idx, (_, row) in enumerate(daily.iterrows(), start=3):
        is_subtotal = bool(row.get("is_subtotal", False))
        row_font = SUBTOTAL_FONT if is_subtotal else DEFAULT_FONT

        for col_idx, header in enumerate(DAILY_HEADERS, start=1):
            value = row[header]
            if header == "Дата" and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            if is_subtotal:
                cell.fill = SUBTOTAL_FILL
            _apply_cell_format(cell, header)

    last_row = len(daily) + 2
    if last_row >= 3:
        _apply_discrepancy_formatting(
            ws,
            cashless_range=f"C3:F{last_row}",
            cashless_diff_col="$E",
            cash_range=f"G3:J{last_row}",
            cash_diff_col="$I",
        )


def _write_totals_sheet(ws, totals: pd.DataFrame) -> None:
    _write_sheet_title(ws, "Итоги сверки по подразделениям", len(TOTALS_HEADERS))
    _write_header_row(ws, TOTALS_HEADERS)

    for row_idx, (_, row) in enumerate(totals.iterrows(), start=3):
        values = [
            row["Подразделение"],
            row["ОФД безнал"],
            row["Реестр карт"],
            row["Разница безнал"],
            row["Статус безнал"],
            row["ОФД наличные"],
            row["Реестр кассовых"],
            row["Разница наличные"],
            row["Статус наличные"],
            row["ОФД кредит"],
            row["Общий статус"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DEFAULT_FONT
            _apply_cell_format(cell, TOTALS_HEADERS[col_idx - 1])

    last_row = len(totals) + 2
    if last_row >= 3:
        _apply_discrepancy_formatting(
            ws,
            cashless_range=f"B3:E{last_row}",
            cashless_diff_col="$D",
            cash_range=f"F3:I{last_row}",
            cash_diff_col="$H",
        )
        _apply_overall_status_formatting(ws, f"K3:K{last_row}")


def _write_discrepancies_sheet(ws, discrepancies: pd.DataFrame) -> None:
    _write_sheet_title(ws, "Только строки с расхождениями", len(DISC_HEADERS))
    _write_header_row(ws, DISC_HEADERS)

    if discrepancies.empty:
        return

    export = discrepancies[
        [
            "Подразделение",
            "Дата",
            "ОФД безнал",
            "Реестр карт",
            "Разница безнал",
            "ОФД наличные",
            "Реестр кассовых",
            "Разница наличные",
            "Источник ОФД",
        ]
    ]
    for row_idx, (_, row) in enumerate(export.iterrows(), start=3):
        for col_idx, header in enumerate(DISC_HEADERS, start=1):
            value = row[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = RED_BOLD_FONT
            _apply_cell_format(cell, header)


def _write_rules_sheet(ws) -> None:
    for col, title in enumerate(("Параметр", "Примененное правило"), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = RULES_HEADER_FONT
        cell.fill = TITLE_FILL

    for row_idx, (param, rule) in enumerate(REPORT_RULES, start=2):
        ws.cell(row=row_idx, column=1, value=param).font = DEFAULT_FONT
        ws.cell(row=row_idx, column=2, value=rule).font = DEFAULT_FONT


def generate_report(result: ReconcileResult, output_path: str) -> str:
    """Сохранить Excel-отчёт в формате эталона и вернуть путь к файлу."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    daily_ws = wb.create_sheet("Сверка по дням")
    totals_ws = wb.create_sheet("Итоги")
    disc_ws = wb.create_sheet("Расхождения")
    rules_ws = wb.create_sheet("Правила")

    _write_daily_sheet(daily_ws, result.daily)
    _write_totals_sheet(totals_ws, result.totals)
    _write_discrepancies_sheet(disc_ws, result.discrepancies)
    _write_rules_sheet(rules_ws)

    _autosize(daily_ws, len(DAILY_HEADERS), len(result.daily) + 2)
    _autosize(totals_ws, len(TOTALS_HEADERS), len(result.totals) + 2)
    _autosize(disc_ws, len(DISC_HEADERS), max(len(result.discrepancies), 1) + 2)
    _autosize(rules_ws, 2, len(REPORT_RULES) + 1)

    wb.save(path)
    return str(path.resolve())
